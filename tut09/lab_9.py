#Imports

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

# Load files

with open('./stud_list.txt', 'r') as f:
    student_list = [line.strip() for line in f]

with open('./dates.txt', 'r') as f:
    lecture_dates = [line.strip() for line in f]

attendance_data = pd.read_csv('/content/input_attendance.csv')

#Valid Date, Roll Number and Name Extraction from CSV

student_dict = {}
attendance_dict = {date: {} for date in lecture_dates}

valid_start = pd.to_datetime('18:00:00').time()
valid_end = pd.to_datetime('20:00:00').time()

def is_valid_time(time):
    return (valid_start <= time <= valid_end)

for index, row in attendance_data.iterrows():
    timestamp_str = row['Timestamp']
    roll_no_name = row['Roll']

    if not isinstance(roll_no_name, str):
        continue

    timestamp = pd.to_datetime(timestamp_str, format='%d/%m/%Y %H:%M:%S')
    roll_no, name = roll_no_name.split(' ', 1)

    if roll_no not in student_dict:
        student_dict[roll_no] = name

    date = timestamp.strftime('%d/%m/%Y')
    time = timestamp.time()

    if date in lecture_dates and is_valid_time(time):
        if roll_no not in attendance_dict[date]:
            attendance_dict[date][roll_no] = 0
        attendance_dict[date][roll_no] += 1

#Excel Sheet Generation and Assigning Titles
wb = Workbook()
ws = wb.active
ws.title = "Attendance Record"

ws.append(["Roll Number", "Name"] + lecture_dates + ["Total Attendance Recorded", "Total Attendance Marked", "Total Attendance Allowed", "Proxy"])

#Define Cell Colours

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Define the border style
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

current_row = 2

# Excel Sheet Processing based on Dict Data of Students
for roll_no, name in student_dict.items():
    total_attendance_recorded = 0
    total_attendance_marked = 0

    ws.cell(row=current_row, column=1, value=roll_no)
    ws.cell(row=current_row, column=2, value=name)

    col_idx = 3
    for date in lecture_dates:
        attendance_count = attendance_dict[date].get(roll_no, 0)
        total_attendance_recorded += attendance_count
        ws.cell(row=current_row, column=col_idx, value=attendance_count)
        # Color cells based on attendance count
        if attendance_count == 2:
            fill = green_fill
            total_attendance_marked += 2
        elif attendance_count == 1:
            fill = yellow_fill
            total_attendance_marked += 1
        elif attendance_count == 0:
            fill = white_fill
        elif attendance_count > 2:
            fill = red_fill  # Mark attendance count > 2 as invalid (Red)

        cell = ws.cell(row=current_row, column=col_idx)
        cell.fill = fill
        cell.border = thin_border
        col_idx += 1

    total_attendance_allowed = 2 * len(lecture_dates)
    proxy = abs(total_attendance_recorded - total_attendance_marked)

    ws.cell(row=current_row, column=col_idx, value=total_attendance_recorded)
    ws.cell(row=current_row, column=col_idx + 1, value=total_attendance_marked)
    ws.cell(row=current_row, column=col_idx + 2, value=total_attendance_allowed)
    ws.cell(row=current_row, column=col_idx + 3, value=proxy)

    current_row += 1

wb.save('./output_excel.xlsx')



